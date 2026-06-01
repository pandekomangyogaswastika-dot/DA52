"""
CV. Dewi Aditya — Production Reports Export (CSV)
GAP #6: Laporan Produksi Komprehensif

Endpoints:
  - GET /api/dewi/reports/production/po/{po_id}/export.csv
  - GET /api/dewi/reports/maklon/client/{client_id}/export.csv
"""
# ruff: noqa: E402
from fastapi import APIRouter, Request, HTTPException
from fastapi.responses import StreamingResponse
from database import get_db
from auth import require_auth
from routes._maklon_adapter import legacy_orders_view as _lmo
import uuid
import re
import csv
import io
from datetime import datetime, timezone

router = APIRouter(prefix="/api/dewi/reports", tags=["dewi-reports"])


def _uid():
    return str(uuid.uuid4())


def _now():
    return datetime.now(timezone.utc)


def _fmt_date(d):
    """Format date untuk CSV"""
    if not d:
        return ""
    if isinstance(d, datetime):
        return d.strftime("%Y-%m-%d")
    return str(d)[:10]


def _fmt_num(n):
    """Format number untuk CSV"""
    try:
        return f"{float(n or 0):,.0f}"
    except Exception:
        return "0"


@router.get("/production/po/{po_id}/export.csv")
async def export_po_report_csv(po_id: str, request: Request):
    """
    Export laporan per PO ke CSV.
    Include: PO info, WO list, stage progress, HPP breakdown.
    """
    await require_auth(request)
    db = get_db()
    
    po = await db.production_pos.find_one({"id": po_id}, {"_id": 0})
    if not po:
        raise HTTPException(404, "Production PO tidak ditemukan.")
    
    # Get WOs via order_id or po_number
    order_id = po.get("order_id")
    wos = []
    if order_id:
        wos = await db.rahaza_work_orders.find(
            {"order_id": order_id, "source": "internal"}, {"_id": 0}
        ).to_list(500)
    
    if not wos:
        po_number = po.get("po_number", "")
        if po_number:
            wos = await db.rahaza_work_orders.find(
                {"order_number_snapshot": {"$regex": re.escape(po_number), "$options": "i"}, "source": "internal"},
                {"_id": 0}
            ).to_list(500)
    
    # Build CSV
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header Section
    writer.writerow(["Laporan Produksi per PO - CV. Dewi Aditya"])
    writer.writerow(["PO Number", po.get("po_number", "")])
    writer.writerow(["Customer", po.get("customer_name", "")])
    writer.writerow(["Status", po.get("status", "")])
    writer.writerow(["Created At", _fmt_date(po.get("created_at"))])
    writer.writerow([])
    
    # WO Table
    writer.writerow(["Work Orders"])
    writer.writerow([
        "WO Number", "Model", "Size", "Qty Order", "Qty Completed", 
        "Status", "Material Cost", "Labor Cost", "Overhead Cost", "Total Cost"
    ])
    
    # Compute HPP for each WO (reuse logic from rahaza_hpp.py)
    # Simplified for export: just fetch basic WO data
    for wo in wos:
        wo_id = wo["id"]
        # Get model/size names
        model = await db.rahaza_models.find_one({"id": wo.get("model_id")}, {"_id": 0, "code": 1, "name": 1})
        size = await db.rahaza_sizes.find_one({"id": wo.get("size_id")}, {"_id": 0, "code": 1})
        
        model_code = model.get("code", "") if model else ""
        size_code = size.get("code", "") if size else ""
        
        # Simplified cost calculation (full logic in rahaza_hpp.py endpoint)
        # For export, we can call the endpoint or do simplified aggregation
        # Here: simplified aggregation
        
        # Material cost
        mi_pipe = [
            {"$match": {"work_order_id": wo_id, "status": "issued"}},
            {"$unwind": "$items"},
            {"$group": {"_id": None, "total": {"$sum": {"$multiply": ["$items.qty_issued", "$items.unit_cost"]}}}}
        ]
        mi_result = await db.rahaza_material_issues.aggregate(mi_pipe).to_list(1)
        material_cost = mi_result[0]["total"] if mi_result else 0
        
        # Labor cost (simplified: count WIP output × default rate)
        wip_count = await db.rahaza_wip_events.count_documents({
            "work_order_id": wo_id, "event_type": "output"
        })
        labor_cost = wip_count * 5000  # fallback rate
        
        # Overhead (simplified: 15% of material+labor)
        overhead_cost = (material_cost + labor_cost) * 0.15
        
        total_cost = material_cost + labor_cost + overhead_cost
        
        # Completed qty
        procs = await db.rahaza_processes.find({"active": True, "is_rework": False}, {"_id": 0}).sort("order_seq", 1).to_list(500)
        completed_qty = 0
        if procs:
            last_proc = procs[-1]
            pipe = [
                {"$match": {"event_type": "output", "work_order_id": wo_id, "process_id": last_proc["id"]}},
                {"$group": {"_id": None, "total": {"$sum": "$qty"}}},
            ]
            res = await db.rahaza_wip_events.aggregate(pipe).to_list(1)
            completed_qty = res[0]["total"] if res else 0
        
        writer.writerow([
            wo.get("wo_number", ""),
            model_code,
            size_code,
            wo.get("qty", 0),
            completed_qty,
            wo.get("status", ""),
            _fmt_num(material_cost),
            _fmt_num(labor_cost),
            _fmt_num(overhead_cost),
            _fmt_num(total_cost),
        ])
    
    writer.writerow([])
    writer.writerow(["Generated at", _now().isoformat()])
    
    # Return CSV
    output.seek(0)
    filename = f"laporan_po_{po.get('po_number', 'unknown').replace('/', '_')}_{datetime.now().strftime('%Y%m%d')}.csv"
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/maklon/client/{client_id}/export.csv")
async def export_maklon_client_report_csv(client_id: str, request: Request):
    """
    Export laporan per Klien Maklon ke CSV.
    Include: Client info, orders list, KPI summary.
    """
    await require_auth(request)
    db = get_db()
    
    client = await db.dewi_maklon_clients.find_one({"id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(404, "Klien maklon tidak ditemukan.")
    
    # Filters
    sp = request.query_params
    date_from = sp.get("date_from")
    date_to = sp.get("date_to")
    
    query = {"client_id": client_id}
    if date_from or date_to:
        query["order_date"] = {}
        if date_from:
            query["order_date"]["$gte"] = date_from
        if date_to:
            query["order_date"]["$lte"] = date_to
    
    orders = await _lmo(db).find(query, {"_id": 0}).to_list(500)
    
    # Build CSV
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header Section
    writer.writerow(["Laporan per Klien Maklon - CV. Dewi Aditya"])
    writer.writerow(["Client Code", client.get("code", "")])
    writer.writerow(["Client Name", client.get("name", "")])
    writer.writerow(["Total Orders", len(orders)])
    writer.writerow([])
    
    # Summary KPIs
    total_qty = 0
    total_revenue = 0
    total_hpp_estimated = 0
    total_hpp_actual = 0
    on_time = 0
    late = 0
    
    for order in orders:
        qty = int(order.get("qty_ordered") or 0)
        price_per_pcs = float(order.get("price_per_pcs") or 0)
        revenue = qty * price_per_pcs
        
        total_qty += qty
        total_revenue += revenue
        
        estimated_hpp = price_per_pcs * qty
        total_hpp_estimated += estimated_hpp
        
        # Simplified actual HPP
        mi_sum_pipe = [
            {"$match": {"order_id": order["id"]}},
            {"$group": {"_id": None, "total": {"$sum": {"$multiply": ["$qty", "$unit_cost"]}}}}
        ]
        mi_result = await db.dewi_maklon_material_issues.aggregate(mi_sum_pipe).to_list(1)
        material_cost = mi_result[0]["total"] if mi_result else 0
        
        labor_est = material_cost * 0.3
        overhead_est = (material_cost + labor_est) * 0.15
        actual_hpp = material_cost + labor_est + overhead_est
        total_hpp_actual += actual_hpp
        
        # On-time check
        if order.get("status") == "completed":
            deadline = order.get("deadline_date")
            completion = order.get("completion_date")
            if completion and deadline and completion <= deadline:
                on_time += 1
            else:
                late += 1
    
    margin = total_revenue - total_hpp_actual
    margin_pct = (margin / total_revenue * 100) if total_revenue > 0 else 0
    on_time_rate = (on_time / (on_time + late) * 100) if (on_time + late) > 0 else 0
    
    writer.writerow(["KPI Summary"])
    writer.writerow(["Total Qty", _fmt_num(total_qty)])
    writer.writerow(["Total Revenue", _fmt_num(total_revenue)])
    writer.writerow(["Total HPP Estimated", _fmt_num(total_hpp_estimated)])
    writer.writerow(["Total HPP Actual", _fmt_num(total_hpp_actual)])
    writer.writerow(["Margin", _fmt_num(margin)])
    writer.writerow(["Margin %", f"{margin_pct:.2f}%"])
    writer.writerow(["On-Time Orders", on_time])
    writer.writerow(["Late Orders", late])
    writer.writerow(["On-Time Rate", f"{on_time_rate:.2f}%"])
    writer.writerow([])
    
    # Orders Table
    writer.writerow(["Order Details"])
    writer.writerow([
        "Order Code", "Product", "Qty", "Status", "Revenue", 
        "HPP Estimated", "HPP Actual", "Margin", "Deadline", "Completion", "On-Time"
    ])
    
    for order in orders:
        qty = int(order.get("qty_ordered") or 0)
        price_per_pcs = float(order.get("price_per_pcs") or 0)
        revenue = qty * price_per_pcs
        
        estimated_hpp = price_per_pcs * qty
        
        # Simplified actual HPP
        mi_sum_pipe = [
            {"$match": {"order_id": order["id"]}},
            {"$group": {"_id": None, "total": {"$sum": {"$multiply": ["$qty", "$unit_cost"]}}}}
        ]
        mi_result = await db.dewi_maklon_material_issues.aggregate(mi_sum_pipe).to_list(1)
        material_cost = mi_result[0]["total"] if mi_result else 0
        
        labor_est = material_cost * 0.3
        overhead_est = (material_cost + labor_est) * 0.15
        actual_hpp = material_cost + labor_est + overhead_est
        
        margin = revenue - actual_hpp
        
        deadline = order.get("deadline_date", "")
        completion = order.get("completion_date", "")
        is_on_time = ""
        if order.get("status") == "completed":
            is_on_time = "Yes" if (completion and deadline and completion <= deadline) else "No"
        
        writer.writerow([
            order.get("order_code", ""),
            order.get("product_name", ""),
            qty,
            order.get("status", ""),
            _fmt_num(revenue),
            _fmt_num(estimated_hpp),
            _fmt_num(actual_hpp),
            _fmt_num(margin),
            deadline,
            completion,
            is_on_time,
        ])
    
    writer.writerow([])
    writer.writerow(["Generated at", _now().isoformat()])
    
    # Return CSV
    output.seek(0)
    filename = f"laporan_client_{client.get('code', 'unknown')}_{datetime.now().strftime('%Y%m%d')}.csv"
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )



# ══════════════════════════════════════════════════════════════════════════════
# Phase 3.4: Excel Export Endpoints
# ══════════════════════════════════════════════════════════════════════════════

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


@router.get("/production/po/{po_id}/export.xlsx")
async def export_po_report_excel(po_id: str, request: Request):
    """
    Export laporan per PO ke Excel dengan formatting rapi.
    """
    await require_auth(request)
    db = get_db()
    
    po = await db.production_pos.find_one({"id": po_id}, {"_id": 0})
    if not po:
        raise HTTPException(404, "Production PO tidak ditemukan.")
    
    # Get WOs
    order_id = po.get("order_id")
    wos = []
    if order_id:
        wos = await db.rahaza_work_orders.find(
            {"order_id": order_id, "source": "internal"}, {"_id": 0}
        ).to_list(500)
    
    if not wos:
        po_number = po.get("po_number", "")
        if po_number:
            wos = await db.rahaza_work_orders.find(
                {"order_number_snapshot": {"$regex": re.escape(po_number), "$options": "i"}, "source": "internal"},
                {"_id": 0}
            ).to_list(500)
    
    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan PO"
    
    # Header styling
    header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:J1')
    ws['A1'] = "LAPORAN PRODUKSI PER PO - CV. Dewi Aditya"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # PO Info
    row = 3
    ws[f'A{row}'] = "PO Number:"
    ws[f'B{row}'] = po.get("po_number", "")
    ws[f'B{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Customer:"
    ws[f'B{row}'] = po.get("customer_name", "")
    
    row += 1
    ws[f'A{row}'] = "Status:"
    ws[f'B{row}'] = po.get("status", "")
    
    row += 1
    ws[f'A{row}'] = "Created:"
    ws[f'B{row}'] = _fmt_date(po.get("created_at"))
    
    # WO Table Header
    row += 2
    headers = ["WO Number", "Model", "Size", "Qty Order", "Qty Done", "Status", "Material Cost", "Labor Cost", "Overhead", "Total Cost"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # WO Data
    for wo in wos:
        row += 1
        wo_id = wo["id"]
        
        # Get model/size
        model = await db.rahaza_models.find_one({"id": wo.get("model_id")}, {"_id": 0, "code": 1})
        size = await db.rahaza_sizes.find_one({"id": wo.get("size_id")}, {"_id": 0, "code": 1})
        
        model_code = model.get("code", "") if model else ""
        size_code = size.get("code", "") if size else ""
        
        # Simplified cost
        mi_pipe = [
            {"$match": {"work_order_id": wo_id, "status": "issued"}},
            {"$unwind": "$items"},
            {"$group": {"_id": None, "total": {"$sum": {"$multiply": ["$items.qty_issued", "$items.unit_cost"]}}}}
        ]
        mi_result = await db.rahaza_material_issues.aggregate(mi_pipe).to_list(1)
        material_cost = mi_result[0]["total"] if mi_result else 0
        
        wip_count = await db.rahaza_wip_events.count_documents({"work_order_id": wo_id, "event_type": "output"})
        labor_cost = wip_count * 5000
        overhead_cost = (material_cost + labor_cost) * 0.15
        total_cost = material_cost + labor_cost + overhead_cost
        
        # Completed qty
        procs = await db.rahaza_processes.find({"active": True, "is_rework": False}, {"_id": 0}).sort("order_seq", 1).to_list(500)
        completed_qty = 0
        if procs:
            last_proc = procs[-1]
            pipe = [
                {"$match": {"event_type": "output", "work_order_id": wo_id, "process_id": last_proc["id"]}},
                {"$group": {"_id": None, "total": {"$sum": "$qty"}}},
            ]
            res = await db.rahaza_wip_events.aggregate(pipe).to_list(1)
            completed_qty = res[0]["total"] if res else 0
        
        # Write row
        data = [
            wo.get("wo_number", ""),
            model_code,
            size_code,
            wo.get("qty", 0),
            completed_qty,
            wo.get("status", ""),
            material_cost,
            labor_cost,
            overhead_cost,
            total_cost,
        ]
        
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = border
            if col_idx >= 7:  # Currency columns
                cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal='right' if col_idx >= 4 else 'left')
    
    # Adjust column widths
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"laporan_po_{po.get('po_number', 'unknown').replace('/', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/maklon/client/{client_id}/export.xlsx")
async def export_maklon_client_report_excel(client_id: str, request: Request):
    """
    Export laporan per Klien Maklon ke Excel dengan formatting rapi.
    """
    await require_auth(request)
    db = get_db()
    
    client = await db.dewi_maklon_clients.find_one({"id": client_id}, {"_id": 0})
    if not client:
        raise HTTPException(404, "Klien maklon tidak ditemukan.")
    
    # Get orders
    sp = request.query_params
    date_from = sp.get("date_from")
    date_to = sp.get("date_to")
    
    query = {"client_id": client_id}
    if date_from or date_to:
        query["order_date"] = {}
        if date_from:
            query["order_date"]["$gte"] = date_from
        if date_to:
            query["order_date"]["$lte"] = date_to
    
    orders = await _lmo(db).find(query, {"_id": 0}).to_list(500)
    
    # Create Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Client"
    
    # Styling
    header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    kpi_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:K1')
    ws['A1'] = "LAPORAN PER KLIEN MAKLON - CV. Dewi Aditya"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Client Info
    row = 3
    ws[f'A{row}'] = "Client Code:"
    ws[f'B{row}'] = client.get("code", "")
    ws[f'B{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Client Name:"
    ws[f'B{row}'] = client.get("name", "")
    ws[f'B{row}'].font = Font(bold=True)
    
    row += 1
    ws[f'A{row}'] = "Total Orders:"
    ws[f'B{row}'] = len(orders)
    
    # KPI Summary
    total_qty = sum(int(o.get("qty_ordered") or 0) for o in orders)
    total_revenue = sum(int(o.get("qty_ordered") or 0) * float(o.get("price_per_pcs") or 0) for o in orders)
    
    row += 2
    ws[f'A{row}'] = "KPI SUMMARY"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    
    row += 1
    kpi_data = [
        ("Total Qty:", total_qty),
        ("Total Revenue:", total_revenue),
    ]
    for label, value in kpi_data:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].fill = kpi_fill
        ws[f'B{row}'].fill = kpi_fill
        if isinstance(value, (int, float)):
            ws[f'B{row}'].number_format = '#,##0'
        row += 1
    
    # Orders Table
    row += 1
    headers = ["Order Code", "Product", "Qty", "Status", "Revenue", "HPP Est", "HPP Actual", "Margin", "Deadline", "Completion", "On-Time"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Order data
    for order in orders:
        row += 1
        qty = int(order.get("qty_ordered") or 0)
        price_per_pcs = float(order.get("price_per_pcs") or 0)
        revenue = qty * price_per_pcs
        estimated_hpp = price_per_pcs * qty
        
        # Simplified actual HPP
        mi_sum_pipe = [
            {"$match": {"order_id": order["id"]}},
            {"$group": {"_id": None, "total": {"$sum": {"$multiply": ["$qty", "$unit_cost"]}}}}
        ]
        mi_result = await db.dewi_maklon_material_issues.aggregate(mi_sum_pipe).to_list(1)
        material_cost = mi_result[0]["total"] if mi_result else 0
        
        labor_est = material_cost * 0.3
        overhead_est = (material_cost + labor_est) * 0.15
        actual_hpp = material_cost + labor_est + overhead_est
        
        margin = revenue - actual_hpp
        
        deadline = order.get("deadline_date", "")
        completion = order.get("completion_date", "")
        is_on_time = ""
        if order.get("status") == "completed":
            is_on_time = "Yes" if (completion and deadline and completion <= deadline) else "No"
        
        data = [
            order.get("order_code", ""),
            order.get("product_name", ""),
            qty,
            order.get("status", ""),
            revenue,
            estimated_hpp,
            actual_hpp,
            margin,
            deadline,
            completion,
            is_on_time,
        ]
        
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = border
            if col_idx in [5, 6, 7, 8]:  # Currency
                cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal='right' if col_idx >= 3 else 'left')
    
    # Column widths
    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14
    
    # Save
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"laporan_client_{client.get('code', 'unknown')}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
