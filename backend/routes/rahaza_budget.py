"""
CV. Dewi Aditya ERP — Budget Module (Session 11)

Endpoints (prefix /api/rahaza/finance/budget):
  GET    /budgets                  — list semua budget header
  POST   /budgets                  — buat budget baru
  GET    /budgets/{id}             — detail budget + items
  PUT    /budgets/{id}             — update header
  DELETE /budgets/{id}             — hapus (hanya draft)
  POST   /budgets/{id}/approve     — approve (draft → approved)
  POST   /budgets/{id}/lock        — lock (approved → locked)
  POST   /budgets/{id}/reopen      — reopen ke draft
  GET    /budgets/{id}/items       — list items budget
  POST   /budgets/{id}/items       — tambah item
  PUT    /budgets/{id}/items/{iid} — update item
  DELETE /budgets/{id}/items/{iid} — hapus item
  GET    /budgets/{id}/variance    — variance report (budgeted vs actual)
  POST   /budgets/import-excel     — import dari Excel
  GET    /budget-summary           — ringkasan semua budget aktif (untuk dashboard)
"""
from fastapi import APIRouter, Request, HTTPException, UploadFile, File
from database import get_db
from auth import require_auth, serialize_doc
from routes.shared import get_pagination_params, paginated_response
from datetime import datetime, timezone, date
from typing import Optional
import uuid
import logging
import io

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/rahaza/finance", tags=["rahaza-budget"])

def _uid(): return str(uuid.uuid4())
def _now(): return datetime.now(timezone.utc)

BUDGET_STATUSES = ["draft", "approved", "locked"]

# ─── BUDGET HEADER ──────────────────────────────────────────────────────────

@router.get("/budgets")
async def list_budgets(request: Request, year: Optional[int] = None, status: Optional[str] = None,
                       cost_center_id: Optional[str] = None):
    await require_auth(request)
    db = get_db()
    q = {}
    if year:
        q["year"] = year
    if status:
        q["status"] = status
    if cost_center_id:
        q["cost_center_id"] = cost_center_id

    use_pagination = "page" in request.query_params
    if use_pagination:
        page, limit, skip = get_pagination_params(request, default_limit=50)
        total = await db.rahaza_budgets.count_documents(q)
        rows = await db.rahaza_budgets.find(q, {"_id": 0}).sort("year", -1).skip(skip).limit(limit).to_list(length=10000)
        return paginated_response(serialize_doc(rows), total, page, limit)
    rows = await db.rahaza_budgets.find(q, {"_id": 0}).sort("year", -1).to_list(200)
    return serialize_doc(rows)


@router.post("/budgets")
async def create_budget(request: Request):
    user = await require_auth(request)
    db = get_db()
    body = await request.json()
    name = (body.get("name") or "").strip()
    year = int(body.get("year") or date.today().year)
    if not name:
        raise HTTPException(400, "Nama budget wajib diisi")
    doc = {
        "id": _uid(),
        "name": name,
        "year": year,
        "period_type": body.get("period_type", "monthly"),  # monthly | quarterly | annual
        "cost_center_id": body.get("cost_center_id") or None,
        "department": (body.get("department") or "").strip(),
        "notes": (body.get("notes") or "").strip(),
        "status": "draft",
        "created_at": _now(),
        "created_by": user.get("email"),
        "approved_at": None,
        "approved_by": None,
        "locked_at": None,
    }
    await db.rahaza_budgets.insert_one(doc)
    # activity log omitted (no blocking path)
    return serialize_doc(doc)


@router.get("/budgets/{bid}")
async def get_budget(bid: str, request: Request):
    await require_auth(request)
    db = get_db()
    budget = await db.rahaza_budgets.find_one({"id": bid}, {"_id": 0})
    if not budget:
        raise HTTPException(404, "Budget tidak ditemukan")
    # Enrich cost center
    if budget.get("cost_center_id"):
        cc = await db.rahaza_cost_centers.find_one({"id": budget["cost_center_id"]}, {"_id": 0})
        budget["cost_center_name"] = (cc or {}).get("name", "")
    # Summary totals
    items = await db.rahaza_budget_items.find({"budget_id": bid}, {"_id": 0}).sort("month", 1).to_list(length=10000)
    budget["total_budgeted"] = round(sum(float(i.get("amount_budgeted") or 0) for i in items), 2)
    budget["item_count"] = len(items)
    return serialize_doc(budget)


@router.put("/budgets/{bid}")
async def update_budget(bid: str, request: Request):
    await require_auth(request)
    db = get_db()
    budget = await db.rahaza_budgets.find_one({"id": bid}, {"_id": 0})
    if not budget:
        raise HTTPException(404, "Budget tidak ditemukan")
    if budget.get("status") == "locked":
        raise HTTPException(409, "Budget sudah dikunci, tidak bisa diedit")
    body = await request.json()
    update = {k: v for k, v in {
        "name": (body.get("name") or "").strip() or None,
        "department": body.get("department"),
        "cost_center_id": body.get("cost_center_id"),
        "notes": body.get("notes"),
        "period_type": body.get("period_type"),
    }.items() if v is not None}
    update["updated_at"] = _now()
    await db.rahaza_budgets.update_one({"id": bid}, {"$set": update})
    out = await db.rahaza_budgets.find_one({"id": bid}, {"_id": 0})
    return serialize_doc(out)


@router.delete("/budgets/{bid}")
async def delete_budget(bid: str, request: Request):
    await require_auth(request)
    db = get_db()
    budget = await db.rahaza_budgets.find_one({"id": bid}, {"_id": 0})
    if not budget:
        raise HTTPException(404, "Budget tidak ditemukan")
    if budget.get("status") != "draft":
        raise HTTPException(409, "Hanya budget berstatus draft yang bisa dihapus")
    await db.rahaza_budget_items.delete_many({"budget_id": bid})
    await db.rahaza_budgets.delete_one({"id": bid})
    return {"ok": True}


@router.post("/budgets/{bid}/approve")
async def approve_budget(bid: str, request: Request):
    user = await require_auth(request)
    db = get_db()
    budget = await db.rahaza_budgets.find_one({"id": bid}, {"_id": 0})
    if not budget:
        raise HTTPException(404, "Budget tidak ditemukan")
    if budget.get("status") != "draft":
        raise HTTPException(409, "Hanya budget draft yang bisa di-approve")
    await db.rahaza_budgets.update_one({"id": bid}, {"$set": {
        "status": "approved",
        "approved_at": _now(),
        "approved_by": user.get("email"),
    }})
    return {"ok": True, "status": "approved"}


@router.post("/budgets/{bid}/lock")
async def lock_budget(bid: str, request: Request):
    user = await require_auth(request)
    db = get_db()
    budget = await db.rahaza_budgets.find_one({"id": bid}, {"_id": 0})
    if not budget:
        raise HTTPException(404, "Budget tidak ditemukan")
    if budget.get("status") != "approved":
        raise HTTPException(409, "Hanya budget approved yang bisa dikunci")
    await db.rahaza_budgets.update_one({"id": bid}, {"$set": {
        "status": "locked",
        "locked_at": _now(),
        "locked_by": user.get("email"),
    }})
    return {"ok": True, "status": "locked"}


@router.post("/budgets/{bid}/reopen")
async def reopen_budget(bid: str, request: Request):
    await require_auth(request)
    db = get_db()
    budget = await db.rahaza_budgets.find_one({"id": bid}, {"_id": 0})
    if not budget:
        raise HTTPException(404, "Budget tidak ditemukan")
    if budget.get("status") not in ("approved", "locked"):
        raise HTTPException(409, "Budget sudah dalam status draft")
    await db.rahaza_budgets.update_one({"id": bid}, {"$set": {
        "status": "draft",
        "approved_at": None, "approved_by": None,
        "locked_at": None,   "locked_by": None,
    }})
    return {"ok": True, "status": "draft"}


# ─── BUDGET ITEMS ────────────────────────────────────────────────────────────

@router.get("/budgets/{bid}/items")
async def list_budget_items(bid: str, request: Request, month: Optional[str] = None):
    await require_auth(request)
    db = get_db()
    q = {"budget_id": bid}
    if month:
        q["month"] = month
    items = await db.rahaza_budget_items.find(q, {"_id": 0}).sort(["month", "account_code"]).to_list(length=10000)
    # Enrich with account name & cost center name
    acc_ids = list({i.get("account_id") for i in items if i.get("account_id")})
    cc_ids  = list({i.get("cost_center_id") for i in items if i.get("cost_center_id")})
    accs = {a["id"]: a for a in await db.rahaza_coa.find({"id": {"$in": acc_ids}}, {"_id": 0}).to_list(500)} if acc_ids else {}
    ccs  = {c["id"]: c for c in await db.rahaza_cost_centers.find({"id": {"$in": cc_ids}}, {"_id": 0}).to_list(500)} if cc_ids else {}
    for i in items:
        acc = accs.get(i.get("account_id")) or {}
        cc  = ccs.get(i.get("cost_center_id")) or {}
        i["account_code"]     = acc.get("code", i.get("account_code", ""))
        i["account_name"]     = acc.get("name", i.get("account_name", ""))
        i["cost_center_name"] = cc.get("name", "")
    return serialize_doc(items)


@router.post("/budgets/{bid}/items")
async def add_budget_item(bid: str, request: Request):
    await require_auth(request)
    db = get_db()
    budget = await db.rahaza_budgets.find_one({"id": bid}, {"_id": 0})
    if not budget:
        raise HTTPException(404, "Budget tidak ditemukan")
    if budget.get("status") == "locked":
        raise HTTPException(409, "Budget terkunci")
    body = await request.json()
    if not body.get("account_id"):
        raise HTTPException(400, "account_id (COA) wajib diisi")
    if not body.get("month"):
        raise HTTPException(400, "month (YYYY-MM) wajib diisi")
    try:
        amt = float(body.get("amount_budgeted") or 0)
    except (ValueError, TypeError):
        raise HTTPException(400, "amount_budgeted harus berupa angka")
    # Enrich account info
    acc = await db.rahaza_coa.find_one({"id": body["account_id"]}, {"_id": 0})
    doc = {
        "id": _uid(),
        "budget_id": bid,
        "account_id": body["account_id"],
        "account_code": (acc or {}).get("code", ""),
        "account_name": (acc or {}).get("name", ""),
        "cost_center_id": body.get("cost_center_id") or budget.get("cost_center_id"),
        "month": body["month"],                   # YYYY-MM
        "amount_budgeted": amt,
        "notes": (body.get("notes") or "").strip(),
        "created_at": _now(),
    }
    await db.rahaza_budget_items.insert_one(doc)
    return serialize_doc(doc)


@router.put("/budgets/{bid}/items/{iid}")
async def update_budget_item(bid: str, iid: str, request: Request):
    await require_auth(request)
    db = get_db()
    budget = await db.rahaza_budgets.find_one({"id": bid}, {"_id": 0})
    if not budget:
        raise HTTPException(404, "Budget tidak ditemukan")
    if budget.get("status") == "locked":
        raise HTTPException(409, "Budget terkunci")
    body = await request.json()
    update = {}
    if "amount_budgeted" in body:
        try:
            update["amount_budgeted"] = float(body["amount_budgeted"])
        except (ValueError, TypeError):
            raise HTTPException(400, "amount_budgeted harus berupa angka")
    if "notes" in body:
        update["notes"] = (body["notes"] or "").strip()
    if "month" in body:
        update["month"] = body["month"]
    if "cost_center_id" in body:
        update["cost_center_id"] = body["cost_center_id"]
    if update:
        update["updated_at"] = _now()
        await db.rahaza_budget_items.update_one({"id": iid, "budget_id": bid}, {"$set": update})
    out = await db.rahaza_budget_items.find_one({"id": iid}, {"_id": 0})
    return serialize_doc(out)


@router.delete("/budgets/{bid}/items/{iid}")
async def delete_budget_item(bid: str, iid: str, request: Request):
    await require_auth(request)
    db = get_db()
    budget = await db.rahaza_budgets.find_one({"id": bid}, {"_id": 0})
    if not budget:
        raise HTTPException(404, "Budget tidak ditemukan")
    if budget.get("status") == "locked":
        raise HTTPException(409, "Budget terkunci")
    await db.rahaza_budget_items.delete_one({"id": iid, "budget_id": bid})
    return {"ok": True}


# ─── VARIANCE REPORT ────────────────────────────────────────────────────────

@router.get("/budgets/{bid}/variance")
async def budget_variance(bid: str, request: Request):
    """
    Budget vs Actual Variance per account per month.
    Actual diambil dari: rahaza_journal_lines (debit side matching account).
    """
    await require_auth(request)
    db = get_db()
    budget = await db.rahaza_budgets.find_one({"id": bid}, {"_id": 0})
    if not budget:
        raise HTTPException(404, "Budget tidak ditemukan")
    budget.get("year", date.today().year)
    # Get all budget items
    items = await db.rahaza_budget_items.find({"budget_id": bid}, {"_id": 0}).sort("month", 1).to_list(length=10000)
    if not items:
        return {"budget": serialize_doc(budget), "rows": [], "summary": {"total_budgeted": 0, "total_actual": 0, "total_variance": 0, "total_variance_pct": 0}}

    # Get unique account_ids and months
    acc_ids = list({i["account_id"] for i in items if i.get("account_id")})
    months  = sorted({i["month"] for i in items if i.get("month")})

    # Fetch actual from journal lines (debit side = expense accounts)
    actual_map = {}  # (account_id, month) → actual_amount
    if acc_ids and months:
        from_m = months[0] + "-01"
        to_m   = months[-1] + "-31"
        jl_rows = await db.rahaza_journal_lines.find(
            {"account_id": {"$in": acc_ids}, "journal_date": {"$gte": from_m, "$lte": to_m}},
            {"_id": 0, "account_id": 1, "journal_date": 1, "debit": 1, "credit": 1}
        ).to_list(length=10000)
        for jl in jl_rows:
            month_key = (jl.get("journal_date") or "")[:7]  # YYYY-MM
            if month_key in months:
                key = (jl["account_id"], month_key)
                debit  = float(jl.get("debit") or 0)
                credit = float(jl.get("credit") or 0)
                actual_map[key] = actual_map.get(key, 0) + (debit - credit)

    # Also check expenses (cost entries)
    if acc_ids and months:
        expenses = await db.rahaza_expenses.find(
            {"account_id": {"$in": acc_ids}}, {"_id": 0, "account_id": 1, "date": 1, "amount": 1}
        ).to_list(length=10000)
        for exp in expenses:
            month_key = (exp.get("date") or "")[:7]
            if month_key in months:
                key = (exp["account_id"], month_key)
                actual_map[key] = actual_map.get(key, 0) + float(exp.get("amount") or 0)

    # Build variance rows
    rows = []
    for item in items:
        acc_id    = item.get("account_id", "")
        month_key = item.get("month", "")
        budgeted  = float(item.get("amount_budgeted") or 0)
        actual    = round(actual_map.get((acc_id, month_key), 0), 2)
        variance  = round(budgeted - actual, 2)
        var_pct   = round((variance / budgeted * 100) if budgeted != 0 else 0, 1)
        rows.append({
            "item_id":        item["id"],
            "account_id":     acc_id,
            "account_code":   item.get("account_code", ""),
            "account_name":   item.get("account_name", ""),
            "cost_center_id": item.get("cost_center_id", ""),
            "month":          month_key,
            "amount_budgeted":budgeted,
            "amount_actual":  actual,
            "variance":       variance,
            "variance_pct":   var_pct,
            "status":         ("over" if variance < 0 else ("under" if variance > 0 else "on_target")),
        })

    total_budgeted = round(sum(r["amount_budgeted"] for r in rows), 2)
    total_actual   = round(sum(r["amount_actual"]   for r in rows), 2)
    total_var      = round(total_budgeted - total_actual, 2)
    total_var_pct  = round((total_var / total_budgeted * 100) if total_budgeted != 0 else 0, 1)
    return {
        "budget": serialize_doc(budget),
        "rows":   rows,
        "summary": {
            "total_budgeted": total_budgeted,
            "total_actual":   total_actual,
            "total_variance": total_var,
            "total_variance_pct": total_var_pct,
        },
    }


# ─── EXCEL IMPORT ────────────────────────────────────────────────────────────

@router.post("/budgets/{bid}/import-excel")
async def import_budget_excel(bid: str, request: Request, file: UploadFile = File(...)):
    """
    Import budget items dari Excel.
    Kolom yang diharapkan: account_code, month (YYYY-MM), amount_budgeted, notes (opsional)
    """
    await require_auth(request)
    db = get_db()
    budget = await db.rahaza_budgets.find_one({"id": bid}, {"_id": 0})
    if not budget:
        raise HTTPException(404, "Budget tidak ditemukan")
    if budget.get("status") == "locked":
        raise HTTPException(409, "Budget terkunci")
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl tidak tersedia")
    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    headers_row = [str(c.value or "").strip().lower() for c in ws[1]]
    required = {"account_code", "month", "amount_budgeted"}
    if not required.issubset(set(headers_row)):
        raise HTTPException(400, f"Kolom wajib: account_code, month, amount_budgeted. Ditemukan: {headers_row}")
    # Build account lookup
    all_accs = await db.rahaza_coa.find({}, {"_id": 0, "id": 1, "code": 1, "name": 1}).to_list(500)
    acc_by_code = {a["code"]: a for a in all_accs}
    imported, skipped = 0, []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rowdata = {headers_row[i]: (row[i] if i < len(row) else None) for i in range(len(headers_row))}
        acc_code = str(rowdata.get("account_code") or "").strip()
        month    = str(rowdata.get("month") or "").strip()
        amt_raw  = rowdata.get("amount_budgeted")
        if not acc_code or not month or amt_raw is None:
            skipped.append({"row": acc_code or "?", "reason": "Kolom kosong"})
            continue
        try:
            amt = float(amt_raw)
        except (ValueError, TypeError):
            skipped.append({"row": acc_code, "reason": f"amount_budgeted tidak valid: {amt_raw}"})
            continue
        # Validate month format YYYY-MM
        if len(month) != 7 or month[4] != "-":
            skipped.append({"row": acc_code, "reason": f"Format month harus YYYY-MM: {month}"})
            continue
        acc = acc_by_code.get(acc_code)
        if not acc:
            skipped.append({"row": acc_code, "reason": f"Akun '{acc_code}' tidak ditemukan di COA"})
            continue
        doc = {
            "id": _uid(),
            "budget_id": bid,
            "account_id": acc["id"],
            "account_code": acc["code"],
            "account_name": acc["name"],
            "cost_center_id": str(rowdata.get("cost_center_code") or "").strip() or budget.get("cost_center_id"),
            "month": month,
            "amount_budgeted": amt,
            "notes": str(rowdata.get("notes") or "").strip(),
            "created_at": _now(),
        }
        await db.rahaza_budget_items.insert_one(doc)
        imported += 1
    return {"imported": imported, "skipped": skipped}


# ─── BUDGET SUMMARY (Dashboard) ─────────────────────────────────────────────

@router.get("/budget-summary")
async def budget_summary(request: Request, year: Optional[int] = None):
    await require_auth(request)
    db = get_db()
    cur_year = year or date.today().year
    budgets = await db.rahaza_budgets.find({"year": cur_year}, {"_id": 0}).to_list(50)
    result = []
    for b in budgets:
        items = await db.rahaza_budget_items.find({"budget_id": b["id"]}, {"_id": 0, "amount_budgeted": 1}).to_list(length=10000)
        total_budgeted = round(sum(float(i.get("amount_budgeted") or 0) for i in items), 2)
        result.append({
            "id":            b["id"],
            "name":          b["name"],
            "year":          b["year"],
            "status":        b["status"],
            "total_budgeted": total_budgeted,
            "item_count":    len(items),
        })
    total_all = round(sum(r["total_budgeted"] for r in result), 2)
    return {"year": cur_year, "budgets": result, "total_budgeted": total_all, "count": len(result)}
