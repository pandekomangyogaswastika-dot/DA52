"""/api/assets/bulk-import/* — preview, execute, template (XLSX), execute-file (multipart)."""
from datetime import date
from fastapi import Request, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
import io
import json as _json

from database import get_db
from auth import require_auth
from ._helpers import router, _uid, _now, _gen_asset_number


@router.post("/bulk-import/preview")
async def bulk_import_preview(request: Request, file: UploadFile = File(...)):
    """Parse CSV/Excel, kembalikan preview baris + kolom untuk column mapping."""
    await require_auth(request)
    content = await file.read()
    try:
        if file.filename.endswith(".csv"):
            import csv
            text = content.decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(io.StringIO(text))
            rows = [dict(r) for r in reader]
        else:
            import pandas as pd
            df = pd.read_excel(io.BytesIO(content), dtype=str).fillna("")
            rows = df.to_dict(orient="records")
    except Exception as e:
        raise HTTPException(400, f"Gagal parse file: {e}")

    if not rows:
        raise HTTPException(400, "File kosong atau tidak ada data.")
    return {
        "columns": list(rows[0].keys()),
        "preview": rows[:5],
        "total_rows": len(rows),
    }


@router.post("/bulk-import/execute")
async def bulk_import_execute(request: Request):
    """Eksekusi bulk import aset dari data yang sudah di-mapping (JSON body)."""
    user = await require_auth(request)
    db = get_db()
    body = await request.json()
    rows = body.get("rows", [])
    category_id = body.get("category_id", "")

    if not rows:
        raise HTTPException(400, "Tidak ada data untuk diimport.")
    if not category_id:
        raise HTTPException(400, "Pilih kategori aset terlebih dahulu.")

    cat = await db.dewi_asset_categories.find_one({"id": category_id})
    if not cat:
        raise HTTPException(404, "Kategori tidak ditemukan.")

    created, errors = [], []
    for i, row in enumerate(rows):
        try:
            name = str(row.get("name", "")).strip()
            if not name:
                errors.append({"row": i + 1, "error": "Nama aset wajib diisi"})
                continue
            purchase_date_str = str(row.get("purchase_date", "")).strip()
            if not purchase_date_str:
                purchase_date_str = date.today().isoformat()
            try:
                purchase_date = date.fromisoformat(purchase_date_str[:10]).isoformat()
            except Exception:
                purchase_date = date.today().isoformat()

            try:
                raw_cost = str(row.get("purchase_cost", "0")).replace(",", "").strip()
                purchase_cost = float(raw_cost) if raw_cost else 0.0
            except Exception:
                purchase_cost = 0.0

            try:
                raw_ul = str(row.get("useful_life_months", "")).strip()
                useful_life_months = int(float(raw_ul)) if raw_ul else (cat.get("useful_life_years", 5) * 12)
            except Exception:
                useful_life_months = cat.get("useful_life_years", 5) * 12

            try:
                raw_rv = str(row.get("residual_value", "0")).replace(",", "").strip()
                residual_value = float(raw_rv) if raw_rv else 0.0
            except Exception:
                residual_value = 0.0

            asset_number = await _gen_asset_number(db, cat["code"])
            doc = {
                "id": _uid(),
                "asset_number": asset_number,
                "name": name,
                "category_id": cat["id"],
                "category_name": cat["name"],
                "purchase_date": purchase_date,
                "purchase_cost": purchase_cost,
                "residual_value": residual_value,
                "useful_life_months": useful_life_months,
                "depreciation_method": cat.get("depr_method", "straight_line"),
                "accumulated_depreciation": 0.0,
                "status": "active",
                "location": str(row.get("location", "")).strip(),
                "department": str(row.get("department", "")).strip(),
                "serial_number": str(row.get("serial_number", "")).strip(),
                "brand": str(row.get("brand", "")).strip(),
                "model": str(row.get("model", "")).strip(),
                "notes": str(row.get("notes", "")).strip(),
                "warranty_expiry_date": str(row.get("warranty_expiry_date", ""))[:10] or None,
                "warranty_provider": str(row.get("warranty_provider", "")).strip(),
                "warranty_terms": str(row.get("warranty_terms", "")).strip(),
                "insurance_policy_number": str(row.get("insurance_policy_number", "")).strip(),
                "insurance_provider": str(row.get("insurance_provider", "")).strip(),
                "insurance_expiry_date": str(row.get("insurance_expiry_date", ""))[:10] or None,
                "insurance_value": float(row.get("insurance_value", 0) or 0),
                "photo_url": None,
                "assigned_to": None,
                "assigned_to_name": "",
                "assigned_at": None,
                "procurement_request_id": None,
                "created_by": user["id"],
                "created_by_name": user.get("name", ""),
                "created_at": _now(),
                "updated_at": _now(),
                "disposed_at": None,
                "journal_purchase_id": None,
            }
            await db.dewi_assets.insert_one(doc)
            created.append(asset_number)
        except Exception as e:
            errors.append({"row": i + 1, "error": str(e)})

    return {
        "ok": True,
        "created_count": len(created),
        "error_count": len(errors),
        "created": created,
        "errors": errors[:10],
    }


@router.get("/bulk-import/template")
async def bulk_import_template(request: Request):
    """Download template Excel untuk bulk import aset."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    await require_auth(request)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Import Aset"
    headers = [
        "name*", "purchase_date*", "purchase_cost*",
        "useful_life_months", "residual_value",
        "serial_number", "brand", "model", "location", "department", "notes",
        "warranty_expiry_date", "warranty_provider", "warranty_terms",
        "insurance_policy_number", "insurance_provider", "insurance_expiry_date", "insurance_value",
    ]
    ws.append(headers)
    ws.append([
        "Laptop Dell XPS 15", "2026-01-01", "15000000",
        "48", "0",
        "SN-DELL-0001", "Dell", "XPS 15", "Kantor Pusat", "IT", "Laptop kerja",
        "2028-01-01", "Dell Support", "On-site 2 tahun",
        "POL-2026-001", "Jasindo", "2027-01-01", "20000000",
    ])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=template_import_aset.xlsx"},
    )


@router.post("/bulk-import/execute-file")
async def bulk_import_execute_file(request: Request, file: UploadFile = File(...)):
    """Eksekusi bulk import dengan upload file + column mapping JSON (multipart form)."""
    user = await require_auth(request)
    db = get_db()

    form = await request.form()
    mapping_json = form.get("mapping", "{}")
    category_id = form.get("category_id", "")
    file_obj = form.get("file")

    try:
        mapping = _json.loads(mapping_json)
    except Exception:
        raise HTTPException(400, "mapping JSON tidak valid")
    if not category_id:
        raise HTTPException(400, "category_id wajib")

    cat = await db.dewi_asset_categories.find_one({"id": category_id})
    if not cat:
        raise HTTPException(404, "Kategori tidak ditemukan.")

    content = await file_obj.read()
    filename = file_obj.filename or ""
    try:
        if filename.endswith(".csv"):
            import csv
            text = content.decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(io.StringIO(text))
            all_rows = [dict(r) for r in reader]
        else:
            import pandas as pd
            df = pd.read_excel(io.BytesIO(content), dtype=str).fillna("")
            all_rows = df.to_dict(orient="records")
    except Exception as e:
        raise HTTPException(400, f"Gagal parse file: {e}")

    if not all_rows:
        raise HTTPException(400, "File kosong atau tidak ada data.")

    def apply_mapping(row):
        out = {}
        for field, col in mapping.items():
            if col and col in row:
                out[field] = row[col]
        return out

    created, errors = [], []
    for i, raw in enumerate(all_rows):
        mapped = apply_mapping(raw)
        try:
            name = str(mapped.get("name", "")).strip()
            if not name:
                errors.append({"row": i + 1, "error": "Nama aset kosong"})
                continue
            purchase_date_str = str(mapped.get("purchase_date", "")).strip()
            try:
                purchase_date = date.fromisoformat(purchase_date_str[:10]).isoformat()
            except Exception:
                purchase_date = date.today().isoformat()
            try:
                purchase_cost = float(str(mapped.get("purchase_cost", "0")).replace(",", "") or "0")
            except Exception:
                purchase_cost = 0.0
            try:
                ul = str(mapped.get("useful_life_months", "")).strip()
                useful_life_months = int(float(ul)) if ul else cat.get("useful_life_years", 5) * 12
            except Exception:
                useful_life_months = cat.get("useful_life_years", 5) * 12
            try:
                rv = str(mapped.get("residual_value", "0")).replace(",", "") or "0"
                residual_value = float(rv)
            except Exception:
                residual_value = 0.0

            asset_number = await _gen_asset_number(db, cat["code"])
            doc = {
                "id": _uid(),
                "asset_number": asset_number,
                "name": name,
                "category_id": cat["id"],
                "category_name": cat["name"],
                "purchase_date": purchase_date,
                "purchase_cost": purchase_cost,
                "residual_value": residual_value,
                "useful_life_months": useful_life_months,
                "depreciation_method": cat.get("depr_method", "straight_line"),
                "accumulated_depreciation": 0.0,
                "status": "active",
                "location": str(mapped.get("location", "")).strip(),
                "department": str(mapped.get("department", "")).strip(),
                "serial_number": str(mapped.get("serial_number", "")).strip(),
                "brand": str(mapped.get("brand", "")).strip(),
                "model": str(mapped.get("model", "")).strip(),
                "notes": str(mapped.get("notes", "")).strip(),
                "warranty_expiry_date": (str(mapped.get("warranty_expiry_date", ""))[:10] or None),
                "warranty_provider": str(mapped.get("warranty_provider", "")).strip(),
                "warranty_terms": str(mapped.get("warranty_terms", "")).strip(),
                "insurance_policy_number": str(mapped.get("insurance_policy_number", "")).strip(),
                "insurance_provider": str(mapped.get("insurance_provider", "")).strip(),
                "insurance_expiry_date": (str(mapped.get("insurance_expiry_date", ""))[:10] or None),
                "insurance_value": float(mapped.get("insurance_value", 0) or 0),
                "photo_url": None,
                "assigned_to": None,
                "assigned_to_name": "",
                "assigned_at": None,
                "procurement_request_id": None,
                "created_by": user["id"],
                "created_by_name": user.get("name", ""),
                "created_at": _now(),
                "updated_at": _now(),
                "disposed_at": None,
                "journal_purchase_id": None,
            }
            await db.dewi_assets.insert_one(doc)
            created.append(asset_number)
        except Exception as e:
            errors.append({"row": i + 1, "error": str(e)})

    return {
        "ok": True,
        "created_count": len(created),
        "error_count": len(errors),
        "created": created[:50],
        "errors": errors[:10],
    }
